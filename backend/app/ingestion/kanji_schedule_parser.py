"""Parser for the kanji weekly schedule (`n3_kanji_merged.xlsx`).

Confirmed format (inspected directly against the real file):
- Sheet "Merged N3 Kanji", header row present:
  [#, Kanji, Source, Study Guide Batch, Difficulty Rank (1=easiest)]
- A second "Notes" sheet documents methodology only; not imported.
"""

from dataclasses import dataclass
from pathlib import Path

import openpyxl

from app.ingestion.base import ParseResult

_SHEET_NAME = "Merged N3 Kanji"


@dataclass
class ParsedKanjiScheduleRow:
    kanji: str
    batch_number: int
    difficulty_rank: int | None
    source_row: int


def parse_kanji_schedule(path: Path) -> ParseResult[ParsedKanjiScheduleRow]:
    result: ParseResult[ParsedKanjiScheduleRow] = ParseResult()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheet = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return result

    for row_idx, row in enumerate(rows[1:], start=2):  # skip header, 1-indexed for display
        if row is None or all(v is None for v in row):
            continue

        try:
            kanji = str(row[1]).strip()
            batch_raw = row[3]
            difficulty_raw = row[4] if len(row) > 4 else None
        except IndexError:
            result.add_warning(row_idx, "row has fewer columns than expected", row)
            continue

        if not kanji:
            result.add_warning(row_idx, "missing kanji character", row)
            continue

        if batch_raw is None:
            result.add_warning(row_idx, "missing batch number", row)
            continue

        try:
            batch_number = int(batch_raw)
        except (TypeError, ValueError):
            result.add_warning(row_idx, f"batch number not an integer: {batch_raw!r}", row)
            continue

        difficulty_rank: int | None
        try:
            difficulty_rank = int(difficulty_raw) if difficulty_raw is not None else None
        except (TypeError, ValueError):
            difficulty_rank = None

        result.rows.append(
            ParsedKanjiScheduleRow(
                kanji=kanji,
                batch_number=batch_number,
                difficulty_rank=difficulty_rank,
                source_row=row_idx,
            )
        )

    return result
